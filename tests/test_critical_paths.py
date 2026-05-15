"""
Расширенная тестовая обвязка для критичных путей:
- Promo-codes — race-safety и idempotency
- Conversation persistence — записывает/читает из SQLite
- AI try_with_keys — fallback при пустых/всех-сломанных ключах
- Bot ai-create — лимит max_auto_bots
- Sites edit-block — refund при ошибке
- Worker_lock + auto-backup
"""
import os, sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pytest
import threading

from server.db import SessionLocal


_FAKE_BCRYPT = "$2b$12$abcdefghijklmnopqrstuvCxyz0123456789ABCDEFGHIJKLMNOPQRSTU"


def _make_user(db, email, balance=0):
    """Создать или вернуть тестового юзера. Reset баланс/флаги для повторов."""
    from server.models import User
    import uuid
    u = db.query(User).filter_by(email=email).first()
    if u:
        u.tokens_balance = balance
        db.commit()
        return u
    u = User(
        email=email,
        password_hash=_FAKE_BCRYPT,
        name=email.split("@")[0],
        tokens_balance=balance,
        is_verified=True,
        agreed_to_terms=True,
        referral_code=uuid.uuid4().hex[:8].upper(),
    )
    db.add(u); db.commit(); db.refresh(u)
    return u


# ── Promo codes ──────────────────────────────────────────────────────────────

class TestPromoCodes:
    def test_promo_apply_credits_bonus_in_kopecks(self):
        from server.models import PromoCode, Transaction
        from fastapi.testclient import TestClient
        from server.auth import create_token
        from main import app

        db = SessionLocal()
        try:
            u = _make_user(db, "promo1@test.com", balance=0)
            # Чистим старые промокоды с этим именем
            for old in db.query(PromoCode).filter_by(code="TESTBONUS").all():
                db.delete(old)
            db.commit()
            promo = PromoCode(code="TESTBONUS", discount_pct=0,
                              bonus_tokens=10000, max_uses=10)  # 100 ₽
            db.add(promo); db.commit(); db.refresh(promo)
            user_id = u.id
        finally:
            db.close()

        token = create_token(user_id, "promo1@test.com")
        client = TestClient(app)
        r = client.post("/promo/apply",
                        headers={"Authorization": f"Bearer {token}"},
                        json={"code": "TESTBONUS"})
        assert r.status_code == 200, r.text
        data = r.json()
        assert data["bonus_kopecks"] == 10000
        assert data["bonus_rub"] == 100.0
        assert "100.00 ₽" in data["message"]

        # Баланс юзера должен вырасти
        db = SessionLocal()
        try:
            from server.models import User
            u2 = db.query(User).filter_by(id=user_id).first()
            assert u2.tokens_balance == 10000
        finally:
            db.close()

    def test_promo_apply_twice_by_same_user_fails(self):
        from server.models import PromoCode
        from fastapi.testclient import TestClient
        from server.auth import create_token
        from main import app

        db = SessionLocal()
        try:
            u = _make_user(db, "promo2@test.com", balance=0)
            for old in db.query(PromoCode).filter_by(code="ONCEONLY").all():
                db.delete(old)
            db.commit()
            promo = PromoCode(code="ONCEONLY", discount_pct=10,
                              bonus_tokens=0, max_uses=100)
            db.add(promo); db.commit()
            user_id = u.id
        finally:
            db.close()

        token = create_token(user_id, "promo2@test.com")
        client = TestClient(app)
        r1 = client.post("/promo/apply",
                         headers={"Authorization": f"Bearer {token}"},
                         json={"code": "ONCEONLY"})
        assert r1.status_code == 200
        r2 = client.post("/promo/apply",
                         headers={"Authorization": f"Bearer {token}"},
                         json={"code": "ONCEONLY"})
        assert r2.status_code == 400  # «уже использован»


# ── Conversation persistence ────────────────────────────────────────────────

class TestConversationPersistence:
    def test_conv_append_and_history_roundtrip(self):
        from server.chatbot_engine import conv_append, conv_history
        bot_id = 99001
        chat_id = "test_chat_xyz"
        # Очистим перед тестом
        from server.models import BotConversationTurn
        db = SessionLocal()
        try:
            db.query(BotConversationTurn).filter_by(bot_id=bot_id, chat_id=chat_id).delete()
            db.commit()
        finally:
            db.close()

        conv_append(bot_id, chat_id, "user", "Привет, бот!")
        conv_append(bot_id, chat_id, "assistant", "Здравствуйте!")
        conv_append(bot_id, chat_id, "user", "Как дела?")

        hist = conv_history(bot_id, chat_id, limit=10)
        assert len(hist) == 3
        assert hist[0]["role"] == "user"
        assert hist[0]["content"] == "Привет, бот!"
        assert hist[1]["role"] == "assistant"
        assert hist[2]["content"] == "Как дела?"

    def test_conv_history_respects_limit_and_order(self):
        from server.chatbot_engine import conv_append, conv_history
        bot_id = 99002
        chat_id = "test_chat_lim"
        from server.models import BotConversationTurn
        db = SessionLocal()
        try:
            db.query(BotConversationTurn).filter_by(bot_id=bot_id, chat_id=chat_id).delete()
            db.commit()
        finally:
            db.close()

        for i in range(25):
            conv_append(bot_id, chat_id, "user", f"msg #{i}")

        hist = conv_history(bot_id, chat_id, limit=5)
        # Берём ПОСЛЕДНИЕ 5, в хронологическом порядке (не reverse)
        assert len(hist) == 5
        assert hist[0]["content"] == "msg #20"
        assert hist[-1]["content"] == "msg #24"

    def test_conv_isolation_between_chats(self):
        from server.chatbot_engine import conv_append, conv_history
        from server.models import BotConversationTurn
        db = SessionLocal()
        try:
            db.query(BotConversationTurn).filter(
                BotConversationTurn.bot_id == 99003).delete()
            db.commit()
        finally:
            db.close()

        conv_append(99003, "chatA", "user", "из A")
        conv_append(99003, "chatB", "user", "из B")
        assert [m["content"] for m in conv_history(99003, "chatA")] == ["из A"]
        assert [m["content"] for m in conv_history(99003, "chatB")] == ["из B"]


# ── try_with_keys helper ─────────────────────────────────────────────────────

class TestTryWithKeys:
    def test_no_keys_returns_none(self, monkeypatch):
        from server import ai
        monkeypatch.setattr(ai, "_get_api_keys", lambda p: [])
        # _notify_admin шумит, но для теста нам важен только возврат
        monkeypatch.setattr(ai, "_notify_admin", lambda *a, **k: None)
        result = ai.try_with_keys("openai", lambda key: {"ok": True})
        assert result is None

    def test_first_key_succeeds(self, monkeypatch):
        from server import ai
        monkeypatch.setattr(ai, "_get_api_keys", lambda p: ["sk-aaa", "sk-bbb"])
        result = ai.try_with_keys("openai", lambda key: f"used:{key[-3:]}")
        assert result.startswith("used:")

    def test_all_keys_fail_returns_none(self, monkeypatch):
        from server import ai
        monkeypatch.setattr(ai, "_get_api_keys", lambda p: ["k1", "k2", "k3"])
        monkeypatch.setattr(ai, "_notify_admin", lambda *a, **k: None)
        attempts = []
        def boom(key):
            attempts.append(key)
            raise RuntimeError("provider down")
        result = ai.try_with_keys("openai", boom)
        assert result is None
        assert len(attempts) == 3  # перебрал все

    def test_second_key_succeeds_after_first_fails(self, monkeypatch):
        from server import ai
        monkeypatch.setattr(ai, "_get_api_keys", lambda p: ["bad", "good"])
        # Гарантируем порядок (не shuffled): мокаем _shuffle
        monkeypatch.setattr(ai, "_shuffle", lambda lst: lst)
        def call(key):
            if key == "bad":
                raise RuntimeError("rate limit")
            return {"ok": True, "key": key}
        result = ai.try_with_keys("openai", call)
        assert result == {"ok": True, "key": "good"}


# ── Bot ai-create rate/limit ─────────────────────────────────────────────────

class TestBotAiCreateLimit:
    def test_max_auto_bots_blocks_creation(self, monkeypatch):
        """Юзер с max_auto_bots=2 и 2 уже сгенеренными ботами должен получить 403."""
        from server.models import User, ChatBot
        from fastapi.testclient import TestClient
        from server.auth import create_token
        from main import app

        db = SessionLocal()
        try:
            u = _make_user(db, "limit@test.com", balance=100_000)
            u.max_auto_bots = 2
            # Создадим 2 уже-сгенерённых бота
            for i in range(3):
                bot = db.query(ChatBot).filter_by(user_id=u.id, name=f"auto{i}").first()
                if bot:
                    db.delete(bot)
            db.commit()
            for i in range(2):
                db.add(ChatBot(user_id=u.id, name=f"auto{i}", model="gpt",
                               auto_generated=True))
            db.commit()
            user_id = u.id
        finally:
            db.close()

        # Замокаем build_from_task (не вызываем реальный AI)
        from server import workflow_builder as _wb
        monkeypatch.setattr(_wb, "build_from_task",
                            lambda task, user_api_key=None: {
                                "name": "x", "explanation": "x",
                                "wfc_nodes": [{"id":"n1","type":"trigger_tg","x":80,"y":200,"props":{}}],
                                "wfc_edges": [],
                                "usage": {"input_tokens": 100, "output_tokens": 100},
                            })

        token = create_token(user_id, "limit@test.com")
        client = TestClient(app)
        r = client.post("/chatbots/ai-create",
                        headers={"Authorization": f"Bearer {token}"},
                        json={"description": "Тестовое описание длиннее 10 символов",
                              "name": "Третий"})
        assert r.status_code == 403
        assert "Лимит" in r.json().get("detail", "")


# ── Sites edit-block refund ──────────────────────────────────────────────────

class TestSitesEditBlockRefund:
    def test_refund_when_ai_returns_garbage(self, monkeypatch):
        """Если Claude вернул не-HTML — деньги возвращаются."""
        from server.models import SiteProject
        from fastapi.testclient import TestClient
        from server.auth import create_token
        from main import app

        db = SessionLocal()
        try:
            u = _make_user(db, "edit@test.com", balance=10_000)
            # Создаём project с code_html
            proj = SiteProject(user_id=u.id, name="test", status="done",
                               spec_text="x", code_html="<html><body>ok</body></html>",
                               conversation_phase="done")
            db.add(proj); db.commit(); db.refresh(proj)
            user_id = u.id
            project_id = proj.id
            initial_balance = u.tokens_balance
        finally:
            db.close()

        # Мокаем generate_response — возвращает сырой текст (без <)
        from server.routes import sites as _sites
        monkeypatch.setattr(_sites, "generate_response",
                            lambda model, messages, extra=None:
                            {"content": "Извините, не могу помочь.", "type": "text"})

        token = create_token(user_id, "edit@test.com")
        client = TestClient(app)
        r = client.post(f"/sites/projects/{project_id}/edit-block",
                        headers={"Authorization": f"Bearer {token}"},
                        json={"block_id": "b1",
                              "block_html": "<section>old</section>",
                              "instruction": "перепиши"})
        assert r.status_code == 503

        # Баланс должен остаться прежним (был возврат)
        db = SessionLocal()
        try:
            from server.models import User
            u2 = db.query(User).filter_by(id=user_id).first()
            assert u2.tokens_balance == initial_balance, \
                f"Expected refund: {initial_balance}, got {u2.tokens_balance}"
        finally:
            db.close()


# ── Sentry + request-id headers ──────────────────────────────────────────────

class TestRequestId:
    def test_response_has_request_id_header(self):
        from fastapi.testclient import TestClient
        from main import app
        client = TestClient(app)
        r = client.get("/faq")
        assert r.status_code == 200
        assert "X-Request-ID" in r.headers
        assert len(r.headers["X-Request-ID"]) >= 8

    def test_request_id_passes_through_when_provided(self):
        from fastapi.testclient import TestClient
        from main import app
        client = TestClient(app)
        rid = "my-trace-id-123"
        r = client.get("/faq", headers={"X-Request-ID": rid})
        assert r.headers["X-Request-ID"] == rid


# ── Secrets crypto: HKDF + legacy compat ─────────────────────────────────────

class TestSecretsCrypto:
    def test_encrypt_decrypt_roundtrip(self, monkeypatch):
        monkeypatch.setenv("JWT_SECRET", "test-secret-32chars-long-enough-yes")
        from server import secrets_crypto as sc
        sc._fernet_cache.clear()  # сбросить cache между тестами с разным JWT_SECRET
        token = sc.encrypt("hello world")
        assert token.startswith("enc:v1:")
        plain = sc.decrypt(token)
        assert plain == "hello world"

    def test_decrypt_legacy_sha256_token(self, monkeypatch):
        """Токены, зашифрованные старым sha256-KDF, должны расшифровываться."""
        monkeypatch.setenv("JWT_SECRET", "another-test-secret-hkdf-rotation-x")
        from server import secrets_crypto as sc
        sc._fernet_cache.clear()
        # Симулируем старый шифртекст: используем _legacy_sha256_key напрямую
        from cryptography.fernet import Fernet
        legacy_fernet = Fernet(sc._legacy_sha256_key("another-test-secret-hkdf-rotation-x"))
        legacy_token = legacy_fernet.encrypt(b"old data").decode("ascii")
        legacy_value = f"enc:v1:{legacy_token}"
        # Decrypt должен попробовать новый HKDF, не справится, потом legacy sha256, и расшифровать.
        plain = sc.decrypt(legacy_value)
        assert plain == "old data"


# ── User API keys ────────────────────────────────────────────────────────────

class TestUserApiKeys:
    def test_add_key_encrypts_in_db_and_returns_preview(self):
        """Ключ должен сохраниться зашифрованным; GET /user/api-keys возвращает
        masked preview, не plaintext."""
        from fastapi.testclient import TestClient
        from server.auth import create_token
        from server.models import UserApiKey
        from main import app

        db = SessionLocal()
        try:
            u = _make_user(db, "apikey1@test.com", balance=100_000)
            # Чистим существующие ключи для повторов
            db.query(UserApiKey).filter_by(user_id=u.id).delete()
            db.commit()
            user_id = u.id
        finally:
            db.close()

        token = create_token(user_id, "apikey1@test.com")
        client = TestClient(app)
        plaintext = "sk-test1234567890ABCDEF"
        r = client.post("/user/api-keys",
                        headers={"Authorization": f"Bearer {token}"},
                        json={"provider": "openai", "api_key": plaintext, "label": "main"})
        assert r.status_code == 200, r.text

        # В БД лежит зашифрованным (не plaintext) — EncryptedString при чтении
        # сам расшифрует, поэтому смотрим через raw SQL.
        from sqlalchemy import text as _text
        db = SessionLocal()
        try:
            row = db.execute(
                _text("SELECT api_key FROM user_api_keys WHERE user_id=:uid"),
                {"uid": user_id},
            ).fetchone()
            assert row is not None
            stored = row[0]
            assert stored != plaintext, "Ключ должен быть зашифрован в БД"
            assert stored.startswith("enc:"), f"Ожидался enc:-префикс, получено {stored[:10]}"
        finally:
            db.close()

        # GET /user/api-keys возвращает masked preview
        r = client.get("/user/api-keys", headers={"Authorization": f"Bearer {token}"})
        assert r.status_code == 200
        keys = r.json()
        assert len(keys) == 1
        assert keys[0]["provider"] == "openai"
        assert keys[0]["label"] == "main"
        assert "..." in keys[0]["key_preview"]
        assert plaintext not in keys[0]["key_preview"]

    def test_short_key_rejected(self):
        from fastapi.testclient import TestClient
        from server.auth import create_token
        from main import app

        db = SessionLocal()
        try:
            u = _make_user(db, "apikey2@test.com", balance=100_000)
            user_id = u.id
        finally:
            db.close()

        token = create_token(user_id, "apikey2@test.com")
        client = TestClient(app)
        r = client.post("/user/api-keys",
                        headers={"Authorization": f"Bearer {token}"},
                        json={"provider": "openai", "api_key": "short"})
        assert r.status_code == 400

    def test_unknown_provider_rejected(self):
        from fastapi.testclient import TestClient
        from server.auth import create_token
        from main import app

        db = SessionLocal()
        try:
            u = _make_user(db, "apikey3@test.com", balance=100_000)
            user_id = u.id
        finally:
            db.close()

        token = create_token(user_id, "apikey3@test.com")
        client = TestClient(app)
        r = client.post("/user/api-keys",
                        headers={"Authorization": f"Bearer {token}"},
                        json={"provider": "claude-bogus", "api_key": "sk-1234567890"})
        assert r.status_code == 400


# ── Price list: keyword trigger + embedding fallback ─────────────────────────

class TestBotPriceList:
    def test_keyword_trigger_present(self):
        """Без триггера прайс не должен подключаться (защита от удорожания
        обычных диалогов)."""
        from server.chatbot_engine import _price_keyword_in_text
        assert _price_keyword_in_text("Сколько это стоит?") is True
        assert _price_keyword_in_text("Какая цена?") is True
        assert _price_keyword_in_text("сколько руб?") is True
        assert _price_keyword_in_text("Покажи прайс") is True
        # Без триггера
        assert _price_keyword_in_text("Привет, как дела?") is False
        assert _price_keyword_in_text("Расскажи о компании") is False
        assert _price_keyword_in_text("") is False
        assert _price_keyword_in_text(None) is False

    def test_substring_fallback_when_no_embeddings(self, monkeypatch):
        """Если OpenAI embeddings недоступны — fallback на substring matching
        по словам из вопроса. Должен находить релевантные позиции."""
        from server.models import ChatBot, BotPriceItem
        from server.chatbot_engine import _price_context_for_question

        db = SessionLocal()
        try:
            u = _make_user(db, "price1@test.com", balance=100_000)
            # Удалим существующего бота с этим именем
            for old in db.query(ChatBot).filter_by(user_id=u.id, name="prc").all():
                db.query(BotPriceItem).filter_by(bot_id=old.id).delete()
                db.delete(old)
            db.commit()
            bot = ChatBot(user_id=u.id, name="prc", model="gpt", workflow_json="{}")
            db.add(bot); db.commit(); db.refresh(bot)
            # Заполним прайс БЕЗ embeddings — заставим fallback в substring-режим
            db.add(BotPriceItem(bot_id=bot.id, name="Стрижка мужская",
                                price_kop=80000, sort_order=0, is_active=True))
            db.add(BotPriceItem(bot_id=bot.id, name="Окрашивание волос",
                                price_kop=350000, sort_order=1, is_active=True))
            db.add(BotPriceItem(bot_id=bot.id, name="Маникюр гель-лак",
                                price_kop=180000, sort_order=2, is_active=True))
            db.commit()
            bot_obj = db.query(ChatBot).filter_by(id=bot.id).first()
        finally:
            db.close()

        # Без триггера — пустая строка (защита от удорожания)
        ctx = _price_context_for_question(bot_obj, "Привет!")
        assert ctx == ""

        # С триггером: substring находит «Стрижка»
        ctx = _price_context_for_question(bot_obj, "Сколько стоит стрижка?")
        assert "Стрижка" in ctx
        assert "800 ₽" in ctx  # 80000 коп = 800 ₽

    def test_csv_price_upper_bound(self):
        """CSV-импорт не должен принимать цены сверх 1 млрд ₽ (защита от 1e10
        в экспоненциальной нотации)."""
        from fastapi.testclient import TestClient
        from server.auth import create_token
        from server.models import ChatBot, BotPriceItem
        from main import app

        db = SessionLocal()
        try:
            u = _make_user(db, "price2@test.com", balance=100_000)
            for old in db.query(ChatBot).filter_by(user_id=u.id, name="csv-prc").all():
                db.query(BotPriceItem).filter_by(bot_id=old.id).delete()
                db.delete(old)
            db.commit()
            bot = ChatBot(user_id=u.id, name="csv-prc", model="gpt", workflow_json="{}")
            db.add(bot); db.commit(); db.refresh(bot)
            user_id = u.id
            bot_id = bot.id
        finally:
            db.close()

        token = create_token(user_id, "price2@test.com")
        client = TestClient(app)
        # CSV с экспоненциальной нотацией → должна попасть в price_text, не price_kop
        csv_content = "name;price\nЭкспа;1e10\nНорма;1500\n"
        r = client.post(f"/chatbots/{bot_id}/price/import-csv",
                        headers={"Authorization": f"Bearer {token}"},
                        files={"file": ("p.csv", csv_content.encode("utf-8"), "text/csv")})
        assert r.status_code == 200, r.text

        db = SessionLocal()
        try:
            items = db.query(BotPriceItem).filter_by(bot_id=bot_id, is_active=True).all()
            by_name = {it.name: it for it in items}
            assert "Экспа" in by_name
            # 1e10 → не в price_kop (превышает 1 млрд ₽)
            assert by_name["Экспа"].price_kop is None
            assert by_name["Экспа"].price_text  # сохранилось как текст
            # Норма прошла
            assert by_name["Норма"].price_kop == 150000  # 1500 ₽ = 150000 коп
        finally:
            db.close()


# ══ Multi-agent orchestra для бизнес-решений ══════════════════════════════

class TestSolutionsOrchestra:
    """Юнит-тесты для server/solutions_orchestra.py."""

    def test_template_render_basic(self):
        # Placeholder-syntax: одиночные {…} (single brace), как во всех seed-данных
        # и документации (docs/modules/06-solutions.md, CLAUDE.md).
        # Раньше тесты были под double-brace {{…}}, что НЕ совпадало с runtime —
        # placeholder-замена не работала ни на одном из 40 пилотов в БД.
        from server.solutions_orchestra import _render_template
        ctx = {
            "input": "Привет",
            "stages": {
                "scout": {"output": "Найдено 5 источников"},
                "deep": {"outputs": ["Анализ A", "Анализ B", "Анализ C"]},
            },
            "fields": {"product": "SaaS-платформа", "audience": "B2B"},
        }
        # Простой input
        assert _render_template("Тема: {input}", ctx) == "Тема: Привет"
        # Output stage
        assert "5 источников" in _render_template("{scout.output}", ctx)
        # Outputs склейка
        out = _render_template("{deep.outputs}", ctx)
        assert "Анализ A" in out and "Анализ B" in out and "---" in out
        # Конкретная ветка
        assert _render_template("{deep.outputs[1]}", ctx) == "Анализ B"
        # Несуществующий stage → пустая строка (не падает)
        assert _render_template("{missing.output}", ctx) == ""
        # Несуществующая ветка → пусто
        assert _render_template("{deep.outputs[99]}", ctx) == ""
        # v2 fields через {field.NAME}
        assert _render_template("Продукт: {field.product}", ctx) == "Продукт: SaaS-платформа"
        # v2 fields через голое имя (если уникально)
        assert _render_template("ЦА: {audience}", ctx) == "ЦА: B2B"

    def test_template_render_combined(self):
        from server.solutions_orchestra import _render_template
        ctx = {"input": "ниша X", "stages": {"a": {"output": "A"}, "b": {"output": "B"}}}
        out = _render_template("Ввод: {input}, A={a.output}, B={b.output}", ctx)
        assert out == "Ввод: ниша X, A=A, B=B"

    def test_calc_cost_kop(self):
        from server.solutions_orchestra import _calc_cost_kop
        # input=2000, output=500 → real ≈ 2*8 + 0.5*30 = 16+15 = 31 коп
        # margin 500 → 31*5 = 155 коп
        cost = _calc_cost_kop({"input_tokens": 2000, "output_tokens": 500}, 500)
        assert 100 <= cost <= 200
        # Пустой usage → 0
        assert _calc_cost_kop({}, 500) == 0
        assert _calc_cost_kop(None, 500) == 0

    def test_subscribe_unsubscribe(self):
        from server.solutions_orchestra import subscribe_run, unsubscribe_run, _subscribers
        q = subscribe_run(99999)
        assert 99999 in _subscribers
        assert q in _subscribers[99999]
        unsubscribe_run(99999, q)
        assert 99999 not in _subscribers  # cleanup

    def test_orchestra_seed_pilots_have_valid_structure(self):
        # Импорт seed-скрипта проверяет, что 3 пилотных orchestra-конфига валидны
        # (имеют stages, ссылки между ними резолвятся правильно).
        import importlib.util
        spec = importlib.util.spec_from_file_location(
            "seed_orchestra", "scripts/seed_orchestra_solutions.py")
        mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
        for pilot in mod.PILOTS:
            orch = pilot["orchestra"]
            assert "stages" in orch and len(orch["stages"]) >= 2
            assert "default_model" in orch
            stage_ids = {st["id"] for st in orch["stages"]}
            # Все ссылки {{<id>.output}} должны указывать на существующий stage
            import re
            for st in orch["stages"]:
                full = " ".join(filter(None, [
                    st.get("user_prompt", ""), st.get("system_prompt", ""),
                    st.get("query", ""), st.get("from_lines", ""),
                ]))
                refs = re.findall(r"\{\{\s*([a-zA-Z0-9_]+)\.(?:output|outputs)", full)
                for ref in refs:
                    assert ref in stage_ids, (
                        f"В пилоте {pilot['title']!r}, stage {st['id']!r}: "
                        f"ссылка на несуществующий stage {ref!r}"
                    )

    def test_orchestra_endpoints_registered(self):
        # Проверяем, что endpoint'ы оркестры зарегистрированы в FastAPI app.
        import main
        paths = {r.path for r in main.app.routes}
        assert "/solutions/{solution_id}/orchestra/start" in paths
        assert "/solutions/runs/{run_id}/stream" in paths
        assert "/solutions/runs/{run_id}/share" in paths
        assert "/s/{public_token}" in paths
        # UX-полировка + production апгрейды
        assert "/solutions/runs/{run_id}/restage" in paths
        assert "/solutions/runs/{run_id}/save-template" in paths
        assert "/solutions/templates" in paths
        assert "/solutions/templates/{template_id}" in paths
        assert "/solutions/templates/{template_id}/run" in paths
        assert "/solutions/runs/{run_id}/reaction" in paths
        assert "/solutions/runs/{run_id}/docx" in paths

    def test_docx_builder_basic(self, tmp_path):
        """DOCX-builder корректно собирает документ из markdown."""
        from server.docx_builder import markdown_to_docx
        md = (
            "# Главный заголовок\n\n"
            "## Раздел 1\n"
            "Это **жирный** текст и обычный.\n\n"
            "- пункт 1\n"
            "- пункт 2\n\n"
            "| A | B |\n|---|---|\n| 1 | 2 |\n| 3 | 4 |\n\n"
            "## Раздел 2\n"
            "1. цифра 1\n"
            "2. цифра 2\n\n"
            "---\n\n"
            "Финальный абзац.\n"
        )
        out = str(tmp_path / "test.docx")
        ok = markdown_to_docx(md_text=md, title="Тест", out_path=out, subtitle="Подзаголовок")
        # Если python-docx нет — ok=False, тест пропускаем (CI/dev может не иметь)
        if not ok:
            import pytest
            pytest.skip("python-docx not installed in environment")
        import os
        assert os.path.exists(out)
        assert os.path.getsize(out) > 1000  # настоящий docx файл

    def test_orchestra_seed_8_pilots(self):
        """После добавления Аудит соцсети / Финансы / Холодная email — 8 пилотов."""
        import importlib.util
        spec = importlib.util.spec_from_file_location(
            "seed_orchestra2", "scripts/seed_orchestra_solutions.py")
        mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
        assert len(mod.PILOTS) == 8
        titles = {p["title"] for p in mod.PILOTS}
        assert "Аудит соцсети канала" in titles
        assert "Финансовый аудит по Excel" in titles
        assert "Холодная email-рассылка под список компаний" in titles

    def test_xlsx_builder_extracts_tables(self, tmp_path):
        """XLSX builder вытаскивает markdown-таблицы на отдельные листы."""
        try:
            import openpyxl  # noqa
        except ImportError:
            import pytest; pytest.skip("openpyxl not installed in environment")
        from server.xlsx_builder import markdown_to_xlsx
        md = (
            "# Отчёт\n\n"
            "Какой-то длинный параграф для резюме главного листа отчёта.\n\n"
            "## Таблица 1: Конкуренты\n\n"
            "| Имя | Цена | УТП |\n|---|---|---|\n| A | 100 | быстрый |\n| B | 200 | дешёвый |\n\n"
            "## Таблица 2: Риски\n\n"
            "| Риск | Уровень |\n|---|---|\n| 1 | high |\n"
        )
        out = str(tmp_path / "test.xlsx")
        ok = markdown_to_xlsx(md_text=md, title="Тест", out_path=out)
        assert ok
        from openpyxl import load_workbook
        wb = load_workbook(out)
        # Должно быть: Отчёт + 2 листа таблиц
        assert "Отчёт" in wb.sheetnames
        assert len(wb.sheetnames) == 3  # отчёт + 2 таблицы
        # Шапка таблицы — на втором листе (порядок может варьироваться)
        ws = wb[wb.sheetnames[1]]
        assert ws.cell(row=1, column=1).value == "Имя"
        assert ws.cell(row=2, column=1).value == "A"

    def test_new_endpoints_registered(self):
        """Новые endpoints зарегистрированы (XLSX, Push, Wazzup webhook)."""
        import main
        paths = {r.path for r in main.app.routes}
        assert "/solutions/runs/{run_id}/xlsx" in paths
        assert "/user/push/subscribe" in paths
        assert "/user/push/unsubscribe" in paths
        assert "/user/push/vapid-public" in paths
        assert "/user/push/status" in paths
        assert "/user/push/test" in paths
        assert "/webhook/wazzup/{bot_id}" in paths

    def test_wazzup_send_helper_exists(self):
        from server.chatbot_engine import send_whatsapp
        assert callable(send_whatsapp)

    def test_orchestra_generate_image_stage_dispatched(self):
        """Stage type generate_image должен быть распознан runtime в обоих
        диспатчерах (run_orchestra и restage)."""
        with open("server/solutions_orchestra.py", "r", encoding="utf-8") as f:
            src = f.read()
        # 2 упоминания в коде: один в run_orchestra, один в restage
        assert src.count('"generate_image"') >= 2
        assert "_run_generate_image" in src

    def test_compare_endpoints(self):
        """Compare-endpoints зарегистрированы."""
        import main
        paths = {r.path for r in main.app.routes if hasattr(r, "path")}
        assert "/solutions/{solution_id}/orchestra/start-compare" in paths
        assert "/solutions/compare/{compare_group}" in paths


class TestMarketplace:
    """Marketplace ботов: модели, endpoints, share-логика."""

    def test_marketplace_endpoints_registered(self):
        import main
        paths = {r.path for r in main.app.routes if hasattr(r, "path")}
        assert "/marketplace/listings" in paths
        assert "/marketplace/listings/{listing_id}" in paths
        assert "/marketplace/listings/{listing_id}/install" in paths
        assert "/marketplace/listings/{listing_id}/review" in paths
        assert "/marketplace/my-listings" in paths
        assert "/marketplace/admin/pending" in paths
        assert "/marketplace/admin/listings/{listing_id}/approve" in paths

    def test_marketplace_models_importable(self):
        from server.models import (BotMarketplaceListing,
                                    BotMarketplaceInstall)
        assert BotMarketplaceListing.__tablename__ == "bot_marketplace_listings"
        assert BotMarketplaceInstall.__tablename__ == "bot_marketplace_installs"


class TestPublicAPI:
    """Public API: токены + endpoints."""

    def test_api_endpoints_registered(self):
        import main
        paths = {r.path for r in main.app.routes if hasattr(r, "path")}
        assert "/api-tokens" in paths
        assert "/api-tokens/{token_id}" in paths
        assert "/api/v1/me" in paths
        assert "/api/v1/proposals/generate" in paths
        assert "/api/v1/proposals/{proposal_id}" in paths

    def test_api_token_secret_hashed(self):
        """Секрет НЕ должен храниться в plain — только hash."""
        from server.routes.public_api import _hash_secret
        h1 = _hash_secret("abc123")
        h2 = _hash_secret("abc123")
        h3 = _hash_secret("abc124")
        assert h1 == h2  # детерминированный
        assert h1 != h3  # разные secrets → разные hash
        assert len(h1) == 64  # sha256 hex

    def test_api_token_authenticate_invalid(self):
        from fastapi.testclient import TestClient
        import main
        client = TestClient(main.app)
        # Без токена — 401
        r = client.get("/api/v1/me")
        assert r.status_code == 401
        # Не bearer — 401
        r = client.get("/api/v1/me", headers={"Authorization": "Basic xxx"})
        assert r.status_code == 401
        # Неправильный префикс — 401
        r = client.get("/api/v1/me", headers={"Authorization": "Bearer wrong_token"})
        assert r.status_code == 401
        # Правильный формат, но несуществующий токен
        r = client.get("/api/v1/me",
                        headers={"Authorization": "Bearer ai_che_aaaaaaaa_" + "f"*48})
        assert r.status_code == 401

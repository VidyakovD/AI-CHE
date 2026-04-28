"""
RAG-knowledge: разбиение файлов на чанки, индексация через OpenAI embeddings,
семантический поиск через cosine similarity.

Использование:
  - add_file(...)       — извлечь текст, разбить на чанки, индексировать
  - retrieve(...)       — найти top-K чанков по cosine similarity к запросу
  - search_kb(...)      — legacy-обёртка для старых нод бота (kb_search)
  - search_file(...)    — поиск файла по name/description (для отправки документа)
  - get_files(...)      — список файлов owner'а
  - delete_file(...)    — удаление файла + всех чанков

Архитектура:
  KnowledgeFile (метадата) → много KnowledgeChunk (текст + embedding).
  owner_type ∈ {"bot", "agent"}, owner_id = bot.id | agent_config.id.
  user_id всегда заполняется — для проверки доступа и storage-биллинга.
"""
import os
import re
import json
import math
import logging
from typing import Iterable, List, Dict, Tuple

from server.db import SessionLocal
from server.models import KnowledgeFile, KnowledgeChunk

log = logging.getLogger("kb")

# OpenAI text-embedding-3-small — 1536-мерный вектор, $0.02 / 1M токенов.
# Этого размера достаточно для большинства RAG-сценариев. Для очень точного
# поиска можно поднять до text-embedding-3-large (3072-мерный, дороже).
EMBED_MODEL = "text-embedding-3-small"
EMBED_DIM = 1536

# Чанкинг: целимся в ~500 токенов, перекрытие 80 (≈ 1-2 предложения).
CHUNK_TARGET_TOKENS = 500
CHUNK_OVERLAP_TOKENS = 80
# Approx 4 chars / token для русского/английского. Жёсткий лимит чанка
# в символах = TARGET * 4 = 2000 символов.
CHARS_PER_TOKEN = 4

# Лимиты на юзера / файл
MAX_FILES_PER_OWNER = 50
MAX_FILE_BYTES = 50 * 1024 * 1024            # 50 МБ
MAX_TOTAL_TEXT_CHARS = 2_000_000             # ~500 страниц A4


# ── Извлечение текста ────────────────────────────────────────────────────────

def _abs_path(rel: str) -> str:
    base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, "..", rel.lstrip("/")) if not os.path.isabs(rel) else rel


def extract_text(file_path: str, mime: str | None = None) -> str:
    """Извлечь текст из файла. Поддержка: PDF, DOCX, XLSX, CSV, TXT, MD, HTML."""
    abs_path = _abs_path(file_path)
    if not os.path.exists(abs_path):
        log.warning(f"[KB] file not found: {file_path}")
        return ""
    ext = os.path.splitext(abs_path)[1].lower()
    try:
        if ext in (".txt", ".md", ".csv", ".json", ".tsv"):
            with open(abs_path, "r", encoding="utf-8", errors="ignore") as f:
                return f.read()[:MAX_TOTAL_TEXT_CHARS]
        if ext == ".html" or ext == ".htm":
            with open(abs_path, "r", encoding="utf-8", errors="ignore") as f:
                html = f.read()
            return _strip_html(html)[:MAX_TOTAL_TEXT_CHARS]
        if ext == ".pdf":
            return _extract_pdf(abs_path)
        if ext == ".docx":
            return _extract_docx(abs_path)
        if ext in (".xlsx", ".xlsm"):
            return _extract_xlsx(abs_path)
        log.warning(f"[KB] unsupported ext: {ext}")
        return ""
    except Exception as e:
        log.error(f"[KB] extract error: {type(e).__name__}: {e}")
        return ""


def _extract_pdf(path: str) -> str:
    from PyPDF2 import PdfReader
    reader = PdfReader(path)
    parts = []
    for i, page in enumerate(reader.pages):
        try:
            t = page.extract_text() or ""
        except Exception:
            t = ""
        if t.strip():
            parts.append(f"[Страница {i+1}]\n{t.strip()}")
    return "\n\n".join(parts)[:MAX_TOTAL_TEXT_CHARS]


def _extract_docx(path: str) -> str:
    import zipfile
    try:
        from defusedxml.ElementTree import parse as _safe_parse  # type: ignore
    except ImportError:
        import xml.etree.ElementTree as ET
        def _safe_parse(f):
            return ET.parse(f)
    parts = []
    with zipfile.ZipFile(path) as z:
        with z.open("word/document.xml") as f:
            tree = _safe_parse(f)
    ns = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
    for el in tree.iter(ns + "t"):
        if el.text:
            parts.append(el.text)
    return "\n".join(parts)[:MAX_TOTAL_TEXT_CHARS]


def _extract_xlsx(path: str) -> str:
    """Excel → плоский текст. Каждый лист → таблица: 'Лист | Колонка1 | Колонка2 | ...'."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        log.warning("[KB] openpyxl not installed — XLSX skipped")
        return ""
    wb = load_workbook(path, data_only=True, read_only=True)
    out_parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        out_parts.append(f"=== Лист: {sheet_name} ===")
        rows_added = 0
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else "" for c in row]
            if not any(cells):
                continue
            out_parts.append(" | ".join(cells))
            rows_added += 1
            if rows_added >= 5000:  # защита от гигантских таблиц
                out_parts.append("[...пропущены остальные строки]")
                break
    wb.close()
    return "\n".join(out_parts)[:MAX_TOTAL_TEXT_CHARS]


def _strip_html(html: str) -> str:
    """Грубое удаление HTML-тегов и схлопывание пробелов."""
    text = re.sub(r"<script[^>]*>.*?</script>", " ", html, flags=re.IGNORECASE | re.DOTALL)
    text = re.sub(r"<style[^>]*>.*?</style>", " ", text, flags=re.IGNORECASE | re.DOTALL)
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


# ── Чанкинг ──────────────────────────────────────────────────────────────────

def _approx_tokens(text: str) -> int:
    return max(1, len(text) // CHARS_PER_TOKEN)


def chunk_text(text: str,
               target_tokens: int = CHUNK_TARGET_TOKENS,
               overlap_tokens: int = CHUNK_OVERLAP_TOKENS) -> List[str]:
    """Режет текст на чанки ~target_tokens с перекрытием overlap_tokens.

    Стратегия: режем по абзацам (\\n\\n), внутри абзаца — по предложениям.
    Для очень длинных сплошных кусков (csv, минифицированный текст без точек)
    используем hard-split по символам.
    """
    if not text or not text.strip():
        return []

    target_chars = target_tokens * CHARS_PER_TOKEN
    overlap_chars = overlap_tokens * CHARS_PER_TOKEN

    # 1. Делим на абзацы (двойной перенос строки)
    paragraphs = re.split(r"\n\s*\n", text)

    chunks: List[str] = []
    buffer = ""

    def _flush(buf: str):
        if buf and buf.strip():
            chunks.append(buf.strip())

    for para in paragraphs:
        para = para.strip()
        if not para:
            continue
        if len(para) > target_chars * 1.5:
            # Слишком длинный абзац — режем по предложениям, потом по символам
            for piece in _split_long(para, target_chars):
                if len(buffer) + len(piece) + 2 <= target_chars:
                    buffer = (buffer + "\n\n" + piece).strip() if buffer else piece
                else:
                    _flush(buffer)
                    buffer = piece
        else:
            if len(buffer) + len(para) + 2 <= target_chars:
                buffer = (buffer + "\n\n" + para).strip() if buffer else para
            else:
                _flush(buffer)
                buffer = para
    _flush(buffer)

    # 2. Перекрытие: к началу каждого следующего чанка добавляем хвост предыдущего.
    if overlap_chars > 0 and len(chunks) > 1:
        with_overlap = [chunks[0]]
        for i in range(1, len(chunks)):
            prev_tail = chunks[i - 1][-overlap_chars:]
            with_overlap.append((prev_tail + "\n\n" + chunks[i]).strip())
        chunks = with_overlap

    return chunks


def _split_long(text: str, target_chars: int) -> List[str]:
    """Длинный кусок без абзацев → режем по предложениям, потом по символам."""
    sentences = re.split(r"(?<=[.!?])\s+", text)
    out: List[str] = []
    buf = ""
    for s in sentences:
        if not s:
            continue
        if len(s) > target_chars:
            # Предложение само огромное (типичная csv-строка) → hard-split
            _flush_buf(out, buf); buf = ""
            for i in range(0, len(s), target_chars):
                out.append(s[i:i + target_chars])
            continue
        if len(buf) + len(s) + 1 <= target_chars:
            buf = (buf + " " + s).strip() if buf else s
        else:
            _flush_buf(out, buf)
            buf = s
    _flush_buf(out, buf)
    return out


def _flush_buf(out: List[str], buf: str):
    if buf and buf.strip():
        out.append(buf.strip())


# ── Embeddings ───────────────────────────────────────────────────────────────

def _embed_batch(texts: List[str]) -> List[List[float]]:
    """Получить embeddings для списка текстов через OpenAI."""
    if not texts:
        return []
    from openai import OpenAI
    from server.ai import _get_api_keys
    keys = _get_api_keys("openai")
    if not keys:
        log.warning("[KB] no OpenAI key — embeddings disabled")
        return [[] for _ in texts]
    cli = OpenAI(api_key=keys[0])
    # Batch до 100 за раз — у OpenAI лимит на размер запроса
    out: List[List[float]] = []
    for i in range(0, len(texts), 100):
        batch = texts[i:i + 100]
        try:
            resp = cli.embeddings.create(model=EMBED_MODEL, input=batch)
            for item in resp.data:
                out.append(list(item.embedding))
        except Exception as e:
            log.error(f"[KB] embeddings batch failed: {type(e).__name__}: {e}")
            out.extend([[] for _ in batch])
    return out


def _embed_one(text: str) -> List[float]:
    out = _embed_batch([text])
    return out[0] if out else []


def _cosine(a: List[float], b: List[float]) -> float:
    if not a or not b or len(a) != len(b):
        return 0.0
    dot = 0.0; na = 0.0; nb = 0.0
    for x, y in zip(a, b):
        dot += x * y; na += x * x; nb += y * y
    if na == 0 or nb == 0:
        return 0.0
    return dot / (math.sqrt(na) * math.sqrt(nb))


# ── Public API: индексация ───────────────────────────────────────────────────

def add_file(*,
             owner_type: str,
             owner_id: int,
             user_id: int,
             name: str,
             path: str,
             mime: str | None = None,
             size: int = 0,
             content_text: str | None = None,
             tags: str = "",
             skip_embeddings: bool = False) -> Dict:
    """
    Индексирует файл для RAG: извлечь текст → разбить на чанки → embedding каждого.

    Если content_text не задан — извлекаем сами через extract_text(path).
    skip_embeddings=True — для тестов / fallback при недоступном OpenAI.

    Возвращает dict с метаданными созданного KnowledgeFile.
    """
    assert owner_type in ("bot", "agent"), f"unknown owner_type: {owner_type}"

    # Извлечение текста
    if content_text is None:
        content_text = extract_text(path, mime)
    text = (content_text or "").strip()

    db = SessionLocal()
    try:
        # Лимиты
        existing_count = (db.query(KnowledgeFile)
                            .filter_by(owner_type=owner_type, owner_id=owner_id)
                            .count())
        if existing_count >= MAX_FILES_PER_OWNER:
            raise ValueError(f"Превышен лимит файлов ({MAX_FILES_PER_OWNER}) для этого "
                             f"{'бота' if owner_type == 'bot' else 'агента'}")

        kf = KnowledgeFile(
            user_id=user_id,
            owner_type=owner_type, owner_id=owner_id,
            bot_id=(owner_id if owner_type == "bot" else None),  # legacy
            name=name[:200], path=path, mime=mime, size=size,
            description=name[:200], tags=tags[:500],
            content_text=text[:50000],  # хвост для совместимости
            indexing_status="indexing",
        )
        db.add(kf); db.commit(); db.refresh(kf)
        kf_id = kf.id
    finally:
        db.close()

    # Чанкинг + embeddings (вне DB-сессии — может быть долго)
    if not text:
        _mark_indexed(kf_id, error="Текст не извлечён", chunks=0)
        return _file_dict(kf_id)

    chunks = chunk_text(text)
    if not chunks:
        _mark_indexed(kf_id, error="Нет чанков", chunks=0)
        return _file_dict(kf_id)

    embeddings: List[List[float]] = []
    if not skip_embeddings:
        try:
            embeddings = _embed_batch(chunks)
        except Exception as e:
            log.error(f"[KB] embedding batch failed: {e}")
            embeddings = [[] for _ in chunks]
    else:
        embeddings = [[] for _ in chunks]

    # Записываем чанки
    db = SessionLocal()
    try:
        for idx, (chunk, emb) in enumerate(zip(chunks, embeddings)):
            db.add(KnowledgeChunk(
                kb_file_id=kf_id,
                chunk_index=idx,
                text=chunk,
                embedding_json=(json.dumps(emb, separators=(",", ":")) if emb else None),
                token_count=_approx_tokens(chunk),
            ))
        db.commit()
    finally:
        db.close()

    _mark_indexed(kf_id, error=None, chunks=len(chunks))

    # Опционально: AI-summary для UI (не блокирующее)
    try:
        _generate_summary(kf_id, text[:8000])
    except Exception as e:
        log.warning(f"[KB] summary skipped: {e}")

    return _file_dict(kf_id)


def _mark_indexed(kf_id: int, error: str | None, chunks: int):
    db = SessionLocal()
    try:
        kf = db.query(KnowledgeFile).filter_by(id=kf_id).first()
        if kf:
            kf.chunk_count = chunks
            kf.indexing_status = "failed" if error and chunks == 0 else "ready"
            kf.indexing_error = error
            db.commit()
    finally:
        db.close()


def _generate_summary(kf_id: int, preview: str):
    """Опционально: короткое описание + теги через AI (для красивого UI)."""
    if not preview.strip():
        return
    from server.ai import generate_response
    prompt = (
        f"Текст:\n{preview}\n\n"
        "Сформируй ответ строго в формате (ничего лишнего):\n"
        "ОПИСАНИЕ: одна строка 10-15 слов\n"
        "ТЕГИ: 3-5 через запятую\n"
        "SUMMARY: 2-3 предложения\n"
    )
    try:
        result = generate_response("gpt", [
            {"role": "system", "content": "Ты индексатор. Отвечай строго по формату."},
            {"role": "user", "content": prompt},
        ], extra={"max_tokens": 200, "temperature": 0})
        raw = result.get("content", "") if isinstance(result, dict) else str(result)
    except Exception as e:
        log.warning(f"[KB] summary AI error: {e}")
        return

    def _xtr(label):
        m = re.search(rf"{label}:\s*(.+?)(?=\n[A-ZА-Я]+:|\Z)", raw, re.DOTALL)
        return m.group(1).strip() if m else ""

    description = _xtr("ОПИСАНИЕ")
    tags = _xtr("ТЕГИ")
    summary = _xtr("SUMMARY")

    db = SessionLocal()
    try:
        kf = db.query(KnowledgeFile).filter_by(id=kf_id).first()
        if not kf:
            return
        if description: kf.description = description[:500]
        if tags and not kf.tags: kf.tags = tags[:500]
        if summary: kf.summary = summary[:2000]
        db.commit()
    finally:
        db.close()


# ── Public API: поиск ────────────────────────────────────────────────────────

def retrieve(owner_type: str, owner_id: int,
             query: str, top: int = 5,
             file_ids: List[int] | None = None) -> List[Dict]:
    """Семантический поиск: top-K чанков по cosine similarity к запросу.

    file_ids — опционально, ограничить поиск конкретными файлами (для агента
    с выборочным доступом к части базы).
    """
    if not query or not query.strip():
        return []
    q_emb = _embed_one(query.strip())
    if not q_emb:
        # Fallback на TF-search
        return _tf_fallback(owner_type, owner_id, query, top, file_ids)

    db = SessionLocal()
    try:
        q = (db.query(KnowledgeChunk, KnowledgeFile)
               .join(KnowledgeFile, KnowledgeChunk.kb_file_id == KnowledgeFile.id)
               .filter(KnowledgeFile.owner_type == owner_type,
                       KnowledgeFile.owner_id == owner_id))
        if file_ids:
            q = q.filter(KnowledgeFile.id.in_(file_ids))
        rows = q.all()
    finally:
        db.close()

    scored = []
    for chunk, kf in rows:
        if not chunk.embedding_json:
            continue
        try:
            v = json.loads(chunk.embedding_json)
        except Exception:
            continue
        score = _cosine(q_emb, v)
        if score > 0.05:  # отбрасываем совсем мусор
            scored.append((score, chunk, kf))
    scored.sort(key=lambda x: x[0], reverse=True)
    return [{
        "chunk_id": c.id, "file_id": kf.id, "file_name": kf.name,
        "chunk_index": c.chunk_index, "text": c.text,
        "score": round(s, 4),
    } for s, c, kf in scored[:top]]


def _tf_fallback(owner_type: str, owner_id: int,
                 query: str, top: int,
                 file_ids: List[int] | None) -> List[Dict]:
    """Простой словарный поиск, если embedding недоступен."""
    db = SessionLocal()
    try:
        q = (db.query(KnowledgeChunk, KnowledgeFile)
               .join(KnowledgeFile, KnowledgeChunk.kb_file_id == KnowledgeFile.id)
               .filter(KnowledgeFile.owner_type == owner_type,
                       KnowledgeFile.owner_id == owner_id))
        if file_ids:
            q = q.filter(KnowledgeFile.id.in_(file_ids))
        rows = q.all()
    finally:
        db.close()

    qw = [w for w in re.split(r"\W+", query.lower()) if len(w) > 2]
    out = []
    for chunk, kf in rows:
        text_l = chunk.text.lower()
        sc = sum(text_l.count(w) for w in qw)
        if sc > 0:
            out.append((sc, chunk, kf))
    out.sort(key=lambda x: x[0], reverse=True)
    return [{
        "chunk_id": c.id, "file_id": kf.id, "file_name": kf.name,
        "chunk_index": c.chunk_index, "text": c.text,
        "score": float(s),
    } for s, c, kf in out[:top]]


def build_context_block(results: List[Dict], max_chars: int = 8000) -> str:
    """Форматирует найденные чанки в текстовый блок для подмешивания в prompt."""
    if not results:
        return ""
    out = ["=== БАЗА ЗНАНИЙ ПОЛЬЗОВАТЕЛЯ ===",
           "Используй факты ниже как авторитетный источник при ответе. "
           "Если в источнике нет — скажи что не знаешь."]
    used = 0
    for i, r in enumerate(results, 1):
        block = f"\n--- [{i}] Файл: {r['file_name']}, фрагмент {r['chunk_index'] + 1} ---\n{r['text']}"
        if used + len(block) > max_chars:
            break
        out.append(block); used += len(block)
    out.append("\n=== КОНЕЦ БАЗЫ ===")
    return "\n".join(out)


# ── Управление ───────────────────────────────────────────────────────────────

def get_files(owner_type: str, owner_id: int) -> List[Dict]:
    db = SessionLocal()
    try:
        rows = (db.query(KnowledgeFile)
                  .filter_by(owner_type=owner_type, owner_id=owner_id)
                  .order_by(KnowledgeFile.created_at.desc()).all())
    finally:
        db.close()
    return [_kf_dict(kf) for kf in rows]


def delete_file(owner_type: str, owner_id: int, file_id: int) -> bool:
    db = SessionLocal()
    try:
        kf = (db.query(KnowledgeFile)
                .filter_by(id=file_id, owner_type=owner_type, owner_id=owner_id)
                .first())
        if not kf:
            return False
        path = kf.path
        # Чанки удалятся каскадом по FK ON DELETE CASCADE
        db.delete(kf)
        db.commit()
    finally:
        db.close()
    # Удаляем файл с диска
    try:
        if path:
            ap = _abs_path(path)
            if os.path.exists(ap):
                os.remove(ap)
    except Exception as e:
        log.warning(f"[KB] file remove failed: {e}")
    return True


def _file_dict(kf_id: int) -> Dict:
    db = SessionLocal()
    try:
        kf = db.query(KnowledgeFile).filter_by(id=kf_id).first()
        if not kf:
            return {}
        return _kf_dict(kf)
    finally:
        db.close()


def _kf_dict(kf: KnowledgeFile) -> Dict:
    return {
        "id": kf.id,
        "name": kf.name,
        "path": kf.path,
        "mime": kf.mime,
        "size": kf.size or 0,
        "description": kf.description or "",
        "tags": kf.tags or "",
        "summary": kf.summary or "",
        "chunk_count": kf.chunk_count or 0,
        "status": kf.indexing_status or "pending",
        "error": kf.indexing_error,
        "created_at": kf.created_at.isoformat() if kf.created_at else None,
    }


# ── Legacy API для существующих нод бота ─────────────────────────────────────

def search_kb(bot_id: int, query: str, top: int = 5) -> List[Dict]:
    """Legacy для chatbot_engine.kb_search/kb_rag нод. Использует новый retrieve()."""
    results = retrieve(owner_type="bot", owner_id=bot_id, query=query, top=top)
    return [{
        "id": r["file_id"], "name": r["file_name"],
        "description": "", "summary": r["text"][:500],
        "facts": "", "score": r["score"],
    } for r in results]


def search_file(bot_id: int, query: str, top: int = 5) -> List[Dict]:
    """Legacy для kb_search_file (ищет файл по name/description/tags)."""
    db = SessionLocal()
    try:
        rows = (db.query(KnowledgeFile)
                  .filter_by(owner_type="bot", owner_id=bot_id).all())
    finally:
        db.close()
    qw = [w for w in re.split(r"\W+", (query or "").lower()) if len(w) > 2]
    scored = []
    for r in rows:
        text = " ".join(filter(None, [r.name, r.description, r.tags])).lower()
        sc = sum(text.count(w) for w in qw)
        if sc > 0:
            scored.append((sc, r))
    scored.sort(key=lambda x: x[0], reverse=True)
    return [{
        "id": r.id, "name": r.name, "path": r.path,
        "description": r.description or "", "score": s,
    } for s, r in scored[:top]]


def get_all_files(bot_id: int) -> List[Dict]:
    """Legacy."""
    return get_files(owner_type="bot", owner_id=bot_id)

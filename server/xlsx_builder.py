"""
Markdown → XLSX экспорт финального отчёта бизнес-решения.

Стратегия:
  - Извлекаем все markdown-таблицы из текста (regex по `|...|...|`).
  - Каждая таблица → отдельный лист с именем по предшествующему H2/H3-заголовку
    (если есть). Иначе «Table N».
  - Первый лист «Отчёт» — TL;DR из 3-5 первых не-табличных абзацев + ссылка
    «полная версия в Markdown / PDF / DOCX».

Если в отчёте нет таблиц вообще — отдаём один лист с текстом построчно.
Не идеально, но юзер получит хоть что-то для дальнейшей работы в Excel.
"""
from __future__ import annotations
import logging
import re

log = logging.getLogger(__name__)

_HEADING_RE = re.compile(r"^(#{1,4})\s+(.+?)\s*#*\s*$")
_TABLE_ROW_RE = re.compile(r"^\s*\|(.+)\|\s*$")
_TABLE_SEP_RE = re.compile(r"^\s*\|?\s*[:\-]+\s*(\|\s*[:\-]+\s*)+\|?\s*$")
_BOLD_RE = re.compile(r"\*\*(.+?)\*\*")
_INLINE_CODE_RE = re.compile(r"`([^`]+)`")


def _strip_md_inline(s: str) -> str:
    """Убираем bold/код-инлайн чтобы в ячейке Excel был чистый текст."""
    s = _BOLD_RE.sub(r"\1", s or "")
    s = _INLINE_CODE_RE.sub(r"\1", s)
    return s.strip()


def _safe_sheet_name(name: str, used: set, idx: int) -> str:
    """Excel: имя листа ≤31 символа, без : \\ / ? * [ ]."""
    base = re.sub(r"[:\\/\?\*\[\]]", "_", (name or "").strip())[:28] or f"Table {idx}"
    candidate = base
    n = 2
    while candidate in used:
        suffix = f"_{n}"
        candidate = base[:31 - len(suffix)] + suffix
        n += 1
        if n > 50:
            candidate = f"Table {idx}_{n}"
            break
    used.add(candidate)
    return candidate


def markdown_to_xlsx(*, md_text: str, title: str, out_path: str,
                      subtitle: str = "") -> bool:
    """Markdown → XLSX. Возвращает True при успехе.

    Структура:
      - Лист «Отчёт» с заголовком, подзаголовком и кратким резюме (3-5 абзацев).
      - Один лист на каждую markdown-таблицу. Имя листа — ближайший
        H2/H3-заголовок над таблицей.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError as e:
        log.warning(f"[xlsx] openpyxl missing: {e}")
        return False

    try:
        wb = Workbook()
        # Главный лист
        ws_main = wb.active
        ws_main.title = "Отчёт"
        ws_main["A1"] = title or "Бизнес-решение"
        ws_main["A1"].font = Font(bold=True, size=16)
        if subtitle:
            ws_main["A2"] = subtitle
            ws_main["A2"].font = Font(italic=True, size=11)
        ws_main.column_dimensions["A"].width = 100

        # Парсинг
        lines = (md_text or "").splitlines()
        n = len(lines)
        # Сборщик резюме (первые 5 не-табличных не-заголовочных параграфов)
        summary: list[str] = []
        # Таблицы: [(name, header[], rows[][])]
        tables: list[tuple[str, list[str], list[list[str]]]] = []
        last_heading = ""

        i = 0
        in_table = False
        cur_header: list[str] = []
        cur_rows: list[list[str]] = []

        def _close_table():
            nonlocal in_table, cur_header, cur_rows
            if in_table and cur_header:
                tables.append((last_heading or f"Table {len(tables)+1}",
                                cur_header, cur_rows))
            in_table = False
            cur_header = []
            cur_rows = []

        while i < n:
            line = lines[i].rstrip()
            mh = _HEADING_RE.match(line)
            if mh:
                _close_table()
                last_heading = _strip_md_inline(mh.group(2))
                i += 1
                continue
            if _TABLE_ROW_RE.match(line):
                if not in_table:
                    cols = [_strip_md_inline(c) for c in line.strip().strip("|").split("|")]
                    # Следующая строка — разделитель?
                    if i + 1 < n and _TABLE_SEP_RE.match(lines[i + 1]):
                        in_table = True
                        cur_header = cols
                        cur_rows = []
                        i += 2
                        continue
                else:
                    cols = [_strip_md_inline(c) for c in line.strip().strip("|").split("|")]
                    cur_rows.append(cols)
                    i += 1
                    continue
            # Не-таблица не-заголовок
            if in_table:
                _close_table()
            if line and len(summary) < 5 and not line.startswith(("|", "#", "-", "*", ">", "```")):
                cleaned = _strip_md_inline(line)
                if len(cleaned) > 30:
                    summary.append(cleaned[:500])
            i += 1
        _close_table()

        # Заполняем главный лист
        row = 4
        if summary:
            ws_main.cell(row=row, column=1, value="Краткое резюме").font = Font(bold=True)
            row += 1
            for p in summary:
                ws_main.cell(row=row, column=1, value=p)
                ws_main.cell(row=row, column=1).alignment = Alignment(wrap_text=True,
                                                                       vertical="top")
                row += 1
        if tables:
            ws_main.cell(row=row + 1, column=1,
                         value=f"Таблицы вынесены на отдельные листы ({len(tables)} шт.)").font = Font(italic=True)
        else:
            ws_main.cell(row=row + 1, column=1,
                         value="Таблиц в отчёте не найдено — см. PDF/DOCX для полного текста").font = Font(italic=True)

        # Каждая таблица — отдельный лист
        used_names: set[str] = {"Отчёт"}
        for idx, (name, header, rows) in enumerate(tables, start=1):
            sheet_name = _safe_sheet_name(name, used_names, idx)
            ws = wb.create_sheet(title=sheet_name)
            cols = max(len(header), max((len(r) for r in rows), default=0))
            # Шапка
            for ci in range(cols):
                cell = ws.cell(row=1, column=ci + 1,
                               value=header[ci] if ci < len(header) else "")
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78",
                                         fill_type="solid")
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
            # Строки
            for ri, r in enumerate(rows, start=2):
                for ci in range(cols):
                    ws.cell(row=ri, column=ci + 1,
                            value=r[ci] if ci < len(r) else "")
                    ws.cell(row=ri, column=ci + 1).alignment = Alignment(
                        wrap_text=True, vertical="top")
            # Авто-ширина (приблизительная)
            for ci in range(cols):
                col_letter = ws.cell(row=1, column=ci + 1).column_letter
                max_len = len(header[ci]) if ci < len(header) else 8
                for r in rows[:30]:
                    if ci < len(r):
                        max_len = max(max_len, min(len(r[ci]), 60))
                ws.column_dimensions[col_letter].width = max(12, min(60, max_len + 2))
            # Зафиксировать шапку
            ws.freeze_panes = "A2"

        wb.save(out_path)
        return True
    except Exception as e:
        log.error(f"[xlsx] generation failed: {type(e).__name__}: {e}")
        return False
